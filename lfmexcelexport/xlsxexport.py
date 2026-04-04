"""
lfmexport/xlsxexport.py

Low-level helpers for writing Last.fm tops and history to Excel.
Uses the same colour palette as plottop/plotop.py (sashamaps.net 20-colour set).
"""

from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel, DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour palette (mirrors plottop/plotop.py selected_colors) ────────────────
# Source: https://sashamaps.net/docs/resources/20-colors/
PALETTE = [
    '#e6194B', '#3cb44b', '#ffe119', '#4363d8', '#f58231',
    '#911eb4', '#42d4f4', '#f032e6', '#bfef45', '#fabed4',
    '#469990', '#dcbeff', '#9A6324', '#fffac8', '#800000',
    '#aaffc3', '#808000', '#ffd8b1', '#000075', '#a9a9a9',
]


def _hex_to_rrggbb(hex_color):
    """Convert '#rrggbb' to plain 'RRGGBB'."""
    return hex_color.lstrip('#').upper()


def _solid_fill_sppr(rrggbb):
    """Return a GraphicalProperties with a solid fill for a chart data point."""
    spPr = GraphicalProperties()
    spPr.solidFill = rrggbb
    return spPr


# ── Sheet styles ──────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F2D3D")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="2E86AB")
SECTION_FONT = Font(bold=True, color="FFFFFF", size=12)
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")
BORDER_SIDE  = Side(style="thin", color="CCCCCC")
CELL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)

CHART_ANCHOR_COL = 5    # column E
CHART_WIDTH      = 22   # cm
CHART_HEIGHT     = 13   # cm
CHART_ROW_SPAN   = 24   # rows reserved vertically per chart


# ── Internal helpers ──────────────────────────────────────────────────────────

def _write_section_title(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font      = SECTION_FONT
    cell.fill      = SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=2)


def _write_table(ws, start_row, headers, rows_data):
    """Write header + data rows. Returns the row index after the last data row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border    = CELL_BORDER
    ws.row_dimensions[start_row].height = 18

    for row_offset, row_data in enumerate(rows_data):
        row_num = start_row + 1 + row_offset
        fill = ALT_FILL if row_offset % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.border = CELL_BORDER
            if fill:
                cell.fill = fill
            if col_idx == len(row_data):
                cell.alignment = Alignment(horizontal="right")

    return start_row + 1 + len(rows_data)


def _add_bar_chart(ws, data_start_row, data_end_row, title, anchor_cell, n_items):
    """
    Insert a native horizontal bar chart.
    Categories read from column A, values from column B.
    Each bar gets a distinct colour from PALETTE.
    """
    chart = BarChart()
    chart.type          = "bar"
    chart.grouping      = "clustered"
    chart.title         = title
    chart.legend        = None
    chart.x_axis.title  = "Plays"
    chart.y_axis.title  = None
    chart.width         = CHART_WIDTH
    chart.height        = CHART_HEIGHT
    chart.style         = 2

    values = Reference(ws, min_col=2, max_col=2,
                       min_row=data_start_row, max_row=data_end_row)
    cats   = Reference(ws, min_col=1, max_col=1,
                       min_row=data_start_row, max_row=data_end_row)

    chart.add_data(values)
    chart.set_categories(cats)
    chart.series[0].title = SeriesLabel(v="Plays")

    series = chart.series[0]
    for i in range(n_items):
        colour = _hex_to_rrggbb(PALETTE[i % len(PALETTE)])
        pt = DataPoint(idx=i)
        pt.spPr = _solid_fill_sppr(colour)
        series.dPt.append(pt)

    ws.add_chart(chart, anchor_cell)


# ── Public API ────────────────────────────────────────────────────────────────

def write_top_section(ws, start_row, section_title, headers, rows_data):
    """
    Write a complete top section (title + table + chart) into ws.

    Parameters
    ----------
    ws            : openpyxl Worksheet
    start_row     : int  – first row to write into
    section_title : str
    headers       : list of 2 strings, e.g. ["Artist", "Plays"]
    rows_data     : list of (label, play_count) tuples

    Returns the next available row after this section.
    """
    _write_section_title(ws, start_row, section_title)
    table_start   = start_row + 1
    table_end_row = _write_table(ws, table_start, headers, rows_data)

    data_start = table_start + 1
    data_end   = table_end_row - 1
    n_items    = max(data_end - data_start + 1, 0)

    if n_items > 0:
        anchor_cell = "%s%d" % (get_column_letter(CHART_ANCHOR_COL), table_start)
        _add_bar_chart(ws, data_start, data_end,
                       section_title, anchor_cell, n_items)

    return max(table_end_row, table_start + CHART_ROW_SPAN) + 2


def set_sheet_column_widths(ws, col_widths_map):
    """Set column widths. col_widths_map: {col_letter: width}"""
    for col, width in col_widths_map.items():
        ws.column_dimensions[col].width = width


def enable_autofilter(ws, min_col, max_col, row=1):
    """Enable Excel auto-filter on a header row."""
    ws.auto_filter.ref = "%s%d:%s%d" % (
        get_column_letter(min_col), row,
        get_column_letter(max_col), row,
    )
