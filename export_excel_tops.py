#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_tops.py

Generates two Excel files from your Last.fm SQLite database:
  - tops_durations.xlsx  : one sheet per time period (last N days + all-time)
  - tops_years.xlsx      : one sheet per year

Usage:
    python export_tops.py
"""

from datetime import datetime

from openpyxl import Workbook

from lfmconf.lfmconf import get_lastfm_conf
from lfmexcelexport.xlsxexport import (
    write_top_section,
    set_sheet_column_widths,
)
from lfmpandas.lfmpandas import (
    retrieve_top_artists_for_duration_as_dataframe,
    retrieve_top_artists_for_year_as_dataframe,
    retrieve_top_albums_for_duration_as_dataframe,
    retrieve_top_albums_for_year_as_dataframe,
    retrieve_top_tracks_for_duration_as_dataframe,
    retrieve_top_tracks_for_year_as_dataframe,
)
from plottop.plotop import Duration

conf        = get_lastfm_conf()
START_YEAR  = conf['lastfm']['service']['startYear']
LIMIT       = conf.get('excelExport', {}).get('limit', 20)


# ── Data helpers ──────────────────────────────────────────────────────────────

def _artists_rows(df):
    return list(df[['Artist', 'PlayCount']].itertuples(index=False, name=None))

def _albums_rows(df):
    return list(df[['ArtistAlbum', 'PlayCount']].itertuples(index=False, name=None))

def _tracks_rows(df):
    return list(df[['ArtistAlbumTrack', 'PlayCount']].itertuples(index=False, name=None))


# ── Sheet builder ─────────────────────────────────────────────────────────────

def _build_sheet(ws, period_value, is_year=False):
    """Fill a worksheet with the three tops for a given period."""
    set_sheet_column_widths(ws, {'A': 45, 'B': 10})

    if is_year:
        df_artists = retrieve_top_artists_for_year_as_dataframe(
            period_value, LIMIT, False)
        df_albums  = retrieve_top_albums_for_year_as_dataframe(
            period_value, LIMIT, False)
        df_tracks  = retrieve_top_tracks_for_year_as_dataframe(
            period_value, LIMIT, False)
    else:
        df_artists = retrieve_top_artists_for_duration_as_dataframe(
            period_value, LIMIT, False)
        df_albums  = retrieve_top_albums_for_duration_as_dataframe(
            period_value, LIMIT, False)
        df_tracks  = retrieve_top_tracks_for_duration_as_dataframe(
            period_value, LIMIT, False)

    current_row = 1

    current_row = write_top_section(
        ws, current_row,
        "Top Artists",
        ["Artist", "Plays"],
        _artists_rows(df_artists),
    )

    current_row = write_top_section(
        ws, current_row,
        "Top Albums",
        ["Album", "Plays"],
        _albums_rows(df_albums),
    )

    write_top_section(
        ws, current_row,
        "Top Tracks",
        ["Track", "Plays"],
        _tracks_rows(df_tracks),
    )


# ── Duration export ───────────────────────────────────────────────────────────

def export_durations(output_path="tops_durations.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    durations = [
        (Duration.WEEK,     "Last 7 days"),
        (Duration.MONTH,    "Last 30 days"),
        (Duration.QUARTER,  "Last 90 days"),
        (Duration.SEMESTER, "Last 180 days"),
        (Duration.YEAR,     "Last 365 days"),
        (Duration.OVERALL,  "All-time"),
    ]

    for duration, sheet_name in durations:
        ws = wb.create_sheet(title=sheet_name)
        _build_sheet(ws, duration.get_value(), is_year=False)
        print("  ✓ %s" % sheet_name)

    wb.save(output_path)
    print("Saved: %s" % output_path)


# ── Years export ──────────────────────────────────────────────────────────────

def export_years(output_path="tops_years.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)

    now   = datetime.now()
    years = range(START_YEAR, now.year + 1)

    for year in reversed(years):
        ws = wb.create_sheet(title=str(year))
        _build_sheet(ws, str(year), is_year=True)
        print("  ✓ %d" % year)

    wb.save(output_path)
    print("Saved: %s" % output_path)


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Exporting tops by duration...")
    export_durations()

    print("\nExporting tops by year...")
    export_years()

    print("\nDone.")
