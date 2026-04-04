#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_history.py

Exports the full listening history from the SQLite database to:
  - history.xlsx : one sheet with all scrobbles, auto-filter on all columns

Usage:
    python export_history.py
"""

import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from lfmdb.lfmdb import select
from lfmexcelexport.xlsxexport import enable_autofilter

# openpyxl rejects strings containing illegal XML characters (control chars)
_ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]')

def _clean(value):
    """Strip illegal XML characters from string values."""
    if isinstance(value, str):
        return _ILLEGAL_CHARS_RE.sub('', value)
    return value

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="1F2D3D")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")
BORDER_SIDE  = Side(style="thin", color="CCCCCC")
CELL_BORDER  = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE,
)

QUERY = """
    SELECT
        id,
        play_date,
        artist_name,
        artist_mbid,
        album_name,
        album_mbid,
        track_name,
        track_mbid,
        track_url,
        play_date_uts
    FROM play
    ORDER BY play_date_uts DESC
"""

HEADERS    = ["ID", "Date", "Artist", "Artist MBID", "Album", "Album MBID",
              "Track", "Track MBID", "Track URL", "Timestamp"]
COL_WIDTHS = {"A": 10, "B": 20, "C": 35, "D": 38, "E": 35,
              "F": 38, "G": 45, "H": 38, "I": 45, "J": 14}


# ── Export ────────────────────────────────────────────────────────────────────

def export_history(output_path="history.xlsx"):
    print("Fetching history from database...")
    rows = select(QUERY)
    total = len(rows)
    print("  %d scrobbles found." % total)

    wb = Workbook()
    ws = wb.active
    ws.title = "History"

    # Column widths
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze the header row
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = CELL_BORDER
    ws.row_dimensions[1].height = 22

    # Data rows
    for row_offset, row_data in enumerate(rows):
        row_num = row_offset + 2
        fill = ALT_FILL if row_offset % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=_clean(value))
            cell.border = CELL_BORDER
            if fill:
                cell.fill = fill
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="left")

        # Progress feedback every 10 000 rows
        if row_num % 10_000 == 0:
            print("  Written %d / %d rows..." % (row_num - 1, total))

    # Auto-filter on all columns
    enable_autofilter(ws, min_col=1, max_col=len(HEADERS), row=1)


    wb.save(output_path)
    print("Saved: %s  (%d scrobbles)" % (output_path, total))


# ── Entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    export_history()
